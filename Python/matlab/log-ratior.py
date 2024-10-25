import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
from sklearn import decomposition

def loadDataSet(filename):
    ex = openpyxl.load_workbook(filename)
    sheet = ex['(+']
    data1 = []
    data2 = []
    for i in range(2, 20):
        res = []
        for j in range(2, 16):
            res.append(sheet.cell(row=i, column=j).value)
        data1.append(res)
    sheet = ex['3#']
    for i in range(2, 51):
        res = []
        for j in range(2, 16):
            res.append(sheet.cell(row=i, column=j).value)
        data2.append(res)
    return data1, data2

def process(data, n, m):
    X = []
    for i in range(0, n):
        res = []
        for j in range(0, m):
            if data[i][j] == 0:
                continue
            else:
                res.append(data[i][j])
        down = 1
        w = []
        for j in range(0, len(res) - 1):
            w.append(res[j] - res[len(res) - 1])
        for j in range(0, len(w)):
            down += np.exp(w[j])
        ptr = []
        for j in range(0, len(w)):
            ptr.append(np.exp(w[j]) / down)
        ptr.append(1 / down)
        x = []
        idx = 0
        for j in range(0, m):
            if data[i][j] == 0:
                x.append(0)
            else:
                x.append(ptr[idx])
                idx += 1
        X.append(x)
    return X

filename = 'Log-ratio#*.xlsx'
data1, data2 = loadDataSet(filename)
print(data1)
print(data2)

data1 = process(data1, len(data1), len(data1[0]))
data2 = process(data2, len(data2), len(data2[0]))

# 创建新的Excel文件并写入数据
wk = openpyxl.Workbook()

# 创建两个新的sheet
sheet1 = wk.create_sheet('(+')
sheet2 = wk.create_sheet('3#')

# 将data1写入sheet1
for i in range(1, len(data1) + 1):
    for j in range(1, len(data1[0]) + 1):
        sheet1.cell(row=i, column=j).value = data1[i - 1][j - 1]

# 将data2写入sheet2
for i in range(1, len(data2) + 1):
    for j in range(1, len(data2[0]) + 1):
        sheet2.cell(row=i, column=j).value = data2[i - 1][j - 1]

# 保存Excel文件
wk.save('../Log-ratio1#*.xlsx')
